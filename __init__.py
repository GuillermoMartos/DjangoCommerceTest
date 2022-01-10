import openpyxl as xl


wb= xl.load_workbook('transactions.xlsx')
sheet= wb['Sheet1']


for row in range(1, sheet.max_row+1):
    for col in range (1, sheet.max_column+1):
        print(sheet.cell(row, col).value)